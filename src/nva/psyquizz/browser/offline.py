# -*- coding: utf-8 -*-
# Copyright (c) 2007-2013 NovaReto GmbH
# cklinger@novareto.de

import cgi
import string
import xlsxwriter
import cStringIO
import shutil
import os
import base64

from backports import tempfile
from cromlech.sqlalchemy import get_session
from datetime import date
from nva.psyquizz.models import IQuizz
from openpyxl import load_workbook
from uvclight import Fields, Page
from uvclight import layer, title, name, menu, context, get_template
from uvclight.auth import require
from zope.component import getUtility
from zope.schema import Choice
from zope.interface.verify import verifyObject
from zope.schema.vocabulary import SimpleTerm, SimpleVocabulary

from ..i18n import _
from ..interfaces import ICompanyRequest
from ..models import ClassSession, Student, CriteriaAnswer
from ..interfaces import QuizzAlreadyCompleted, QuizzClosed


CHUNK = 4096


class OfflineQuizz(Page):
    context(ClassSession)
    layer(ICompanyRequest)
    require('manage.company')

    def get_criterias_questions(self):
        for criteria in self.context.course.criterias:
            values = SimpleVocabulary([
                    SimpleTerm(value=c.strip(), token=idx, title=c.strip())
                    for idx, c in enumerate(criteria.items.split('\n'), 1)
                    if c.strip()])

            criteria_field = Choice(
                __name__='criteria_%s' % criteria.id,
                title=criteria.title,
                description=u"Wählen Sie das Zutreffende aus.",
                vocabulary=values,
                required=True,
            )
            yield criteria_field

    def get_quizz_questions(self):
        quizz = getUtility(IQuizz, self.context.quizz_type)
        fields = Fields(quizz.__schema__)
        fields.sort(key=lambda c: c.interface[c.identifier].order)
        for field in fields:
            yield field, field.identifier, field.description
        
        questions_text = self.context.course.extra_questions
        if questions_text:
            for idx, q in enumerate(questions_text.strip().split('\n')):
                yield None, 'extra%s' % idx, q

        criterias_questions = Fields(*list(self.get_criterias_questions()))
        for criteria_field in criterias_questions:
            yield (
                criteria_field,
                criteria_field.identifier,
                criteria_field.description)

                
    def generateXLSX(self, folder, filename="ouput.xlsx"):
        filepath = os.path.join(folder, filename)
        workbook = xlsxwriter.Workbook(filepath)

        worksheet = workbook.add_worksheet()
        
        # Add a format for the header cells.
        header_format = workbook.add_format({
            'border': 1,
            'bg_color': '#C6EFCE',
            'bold': True,
            'text_wrap': True,
            'valign': 'vcenter',
            'indent': 1,
            'locked': 1,
        })

        uuid_format = workbook.add_format({
            'border': 1,
            'color': '#8f0000',
            'bold': True,
            'text_wrap': False,
            'valign': 'vcenter',
            'indent': 0,
            'locked': 1,
        })

        question_format = workbook.add_format({
            'border': 0,
            'color': '#000000',
            'bold': True,
            'text_wrap': False,
            'valign': 'vcenter',
            'indent': 0,
            'locked': 1,
        })

        # Set up layout of the worksheet.
        worksheet.set_column('A:A', 20)
        worksheet.set_column('B:B', 135)

        # Assign styles to headers
        worksheet.write('A1', 'Identifier', header_format)
        worksheet.write('B1', 'Question', header_format)

        worksheet.write('A2', 'UUID', question_format)
        worksheet.write('B2', 'UUID', question_format)

        # Answer columns with their uuids
        for column in string.ascii_uppercase[2:]:
            worksheet.set_column('%s:%s' % (column, column), 40)
            worksheet.write('%s1' % column, 'Answer', header_format)
            uuid = Student.generate_access()
            worksheet.write('%s2' % column, uuid, uuid_format)

        # Filling up lines
        idx = 2
        for field, id, title in self.get_quizz_questions():
            worksheet.write(idx, 0, id, question_format)
            worksheet.write(idx, 1, title, question_format)
            if field is not None:
                vocabulary = getattr(field, 'source', None)
                if vocabulary is not None:
                    # flattened = [i.title for i in vocabulary]
                    # They want to try Values instead of the Title to have a
                    # faster typing of the results
                    flattened = [i.value for i in vocabulary]
                    for column in string.ascii_uppercase[2:]:
                        worksheet.data_validation(
                            '%s%d' % (column, idx + 1),
                            {'validate': 'list',
                             'source': flattened})

            idx += 1
        workbook.close()
        return filepath

    def render(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            filepath = self.generateXLSX(temp_dir)
            output = cStringIO.StringIO()
            with open(filepath, 'rb') as fd:
                shutil.copyfileobj(fd, output)
            output.seek(0)
        return output

    def make_response(self, result):
        response = self.responseFactory()
        response.headers['Content-Type'] = (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response.headers['Content-Disposition'] = (
            u'attachment; filename="offline.xlsx"')

        def filebody(r):
            data = r.read(CHUNK)
            while data:
                yield data
                data = r.read(CHUNK)

        response.app_iter = filebody(result)
        return response



class UploadOfflineQuizz(Page):
    context(ClassSession)
    layer(ICompanyRequest)
    require('manage.company')

    template = get_template('upload.cpt', __file__)

    def get_quizz(self):
        quizz = getUtility(IQuizz, self.context.quizz_type)
        fields = Fields(quizz.__schema__)
        fields.sort(key=lambda c: c.interface[c.identifier].order)
        return quizz, fields

    def read_xls(self, fields, finput):
        answer_objects = {}
        
        wb = load_workbook(filename=finput, read_only=True)
        sheet = wb['Sheet1']
        rows = sheet.rows
        rows.next() # passing headers
        for row in rows:
             identifier, title = [c.value for c in row[0:2]]
             answers = [c.value for c in row[2:]]
             errors = set()
             print identifier
             if identifier == 'UUID':
                 for idx, answer in enumerate(answers):
                     answer_object = answer_objects.setdefault(idx, {})
                     answer_object[identifier] = answer
             elif identifier == None:
                 continue
             else:
                field = fields.get(identifier)
                if field is not None:
                    for idx, answer in enumerate(answers):
                        if answer is not None and idx not in errors:
                            answer_object = answer_objects.setdefault(idx, {})
                            if not field.source:
                                answer_object[identifier] = answer
                            else:
                                try:
                                    #import pdb; pdb.set_trace()
                                    #token = base64.b64encode(
                                    #    answer.encode('utf8'))
                                    value = field.source.getTerm(int(answer))
                                    answer_object[identifier] = value.value
                                except:
                                    # Value doesn't exist, HELP !
                                    raise
                        else:
                            if idx not in errors:
                                if idx in answer_objects:
                                    del answer_objects[idx]
                                errors.add(idx)
                            continue
                else:
                    # this is a big error, handle me
                    raise NotImplementedError
 
        return answer_objects

    def update(self):

        #if date.today() > self.context.enddate:
        #    raise QuizzClosed(self)

        if self.request.method.upper() == 'POST':
            if self.request.form.get('submit') == u'Upload':
                f = self.request.form.get('file')
                if isinstance(f, cgi.FieldStorage):

                    quizz, fields = self.get_quizz()
                    session = get_session('school')
                    
                    with tempfile.TemporaryDirectory() as temp_dir:
                        path = os.path.join(temp_dir, 'upload.xlsx')
                        with open(path, 'wb+') as fd:
                            shutil.copyfileobj(f.file, fd)
                        answers = self.read_xls(fields, path)

                    # create answers
                    for idx, answer in answers.items():

                        if 'UUID' not in answer:
                            uuid = Student.generate_access()
                        else:
                            uuid = answer.pop('UUID')

                        student = Student(
                            anonymous=True,
                            access=uuid,
                            completion_date=date.today(),
                            company_id=self.context.course.company_id,
                            session_id=self.context.id,
                            course=self.context.course,
                            quizz_type=self.context.course.quizz_type)

                        criterias = []
                        for key in answer:
                            if key.startswith('criteria_'):
                                cid = key.split('_', 1)[1]
                                value = answer.pop(key)
                                criteria_answer = CriteriaAnswer(
                                    criteria_id=cid,
                                    student_id=student.access,
                                    session_id=self.context.id,
                                    answer=value,
                                )
                                criterias.append(criteria_answer)

                        answer = quizz(
                            student_id=student.access,
                            course_id=student.course_id,
                            session_id=student.session_id,
                            company_id=student.company_id,
                            completion_date = student.completion_date,
                            **answer)

                        assert verifyObject(quizz.__schema__, answer)
                        session.add(student)
                        session.add(answer)
                        for ca in criterias:
                            session.add(ca)
                        self.flash('Ihre Vorlage wurde Erfolgreich importiert')
                    self.redirect(self.application_url())
                        
                else:
                    # error, malformed form
                    raise NotImplementedError
            else:
                # error, wrong action
                raise NotImplementedError

                        
